import numpy as np
import pandas as pd
import openpyxl
import json
import io
import base64

class SampleMetadataUploadChecker:
    '''
    for all of these, if they return true, there is a problem
    '''

    def __init__(self,content_string,valid_columns):
        self.content_string=content_string
        self.total_header_set=valid_columns
        
    def create_workbook(self):
        decoded=base64.b64decode(self.content_string)
        try:
            self.workbook=openpyxl.load_workbook(io.BytesIO(decoded))
            return False
        except:
            return 'Not a valid Excel File'

    def lacks_sheetname(self):
        if ('sample_sheet' not in self.workbook.sheetnames):
            return 'sheet named \"sample_sheet\" not found'
        if ('author_metadata' not in self.workbook.sheetnames):
            return 'sheet named \"sample_sheet\" not found'
                        
        else:
            return False

    def create_dataframe(self):
        decoded=base64.b64decode(self.content_string)
        self.dataframe=pd.read_excel(
            io.BytesIO(decoded),
            sheet_name='sample_sheet',
            index_col=0
            #skiprows=1
        )

    def headers_malformed(self):
        presented_columns={temp_col.split('.')[0] for temp_col in self.dataframe.columns}
        malformed_columns=presented_columns.difference(self.total_header_set)
        if len(malformed_columns)>0:
            return 'The following illegal columns were found: '+', '.join(malformed_columns)
        else:
            return False

    def contains_underscore(self):
        if np.any(self.dataframe.stack().astype(str).str.contains('_').values):
            return 'Metadata cannot contain the character \"_\"'
        else:
            return False

    def contains_no_sample_rows(self):
        if len(self.dataframe.index)==0:
            return 'Form must contain at least one sample row'
        else:
            return False

if __name__=="__main__":

    sample_excel_address='/home/rictuar/Downloads/PRESENTATION.xlsx'

    my_SampleMetadataUploadChecker=SampleMetadataUploadChecker(
        content_string,
        header_address
    )
    my_SampleMetadataUploadChecker.create_workbook()
    my_SampleMetadataUploadChecker.lacks_sheetname()
    my_SampleMetadataUploadChecker.create_dataframe()
    my_SampleMetadataUploadChecker.headers_malformed()
    my_SampleMetadataUploadChecker.contains_underscore()
    my_SampleMetadataUploadChecker.contains_no_sample_rows()